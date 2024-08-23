import pytest
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from openpyxl import load_workbook
from datetime import datetime, timedelta

excel_file = "E:\\Offline_Batch_19\\Projects\\Automation_pnt19\\Project1-Fight_cost\\test_data.xlsx"


# Function to read data from Excel
def get_test_data_from_excel():
    df = pd.read_excel(excel_file)
    return df


# Function to write data to Excel
def write_to_excel(row, calculated_fare, result):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.cell(row=row, column=6).value = calculated_fare
    sheet.cell(row=row, column=7).value = result
    workbook.save(excel_file)


@pytest.fixture
def setup(scope="session"):
    driver = webdriver.Firefox()
    driver.get("https://muntasir101.github.io/Ticket-Fare/")
    yield driver
    driver.quit()


# Read test data from Excel and pass it to the test function
@pytest.mark.parametrize("row, distance_value, ticket_class, number_of_tickets, extra_baggage,expected_cost",
                         get_test_data_from_excel().itertuples(index=True, name=None))
def test_calculate_fare(setup, row, distance_value, ticket_class, number_of_tickets, extra_baggage, expected_cost):
    # Set distance_value
    distance = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "input#distance"))
    )
    distance.send_keys(str(distance_value))

    # Set date
    departure_date = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "input#departureDate"))
    )

    # next day
    current_date_plus_1 = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    departure_date.send_keys(current_date_plus_1)

    # Set class
    service_class = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "select#serviceClass"))
    )
    service_class_options = Select(service_class)
    service_class_options.select_by_value(ticket_class)

    # Set number of tickets
    number_of_tickets_field = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "input#numberOfTickets"))
    )
    number_of_tickets_field.send_keys(str(number_of_tickets))

    # Set extra baggage
    extra_baggage_field = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "input#extraBaggage"))
    )
    extra_baggage_field.send_keys(str(extra_baggage))

    # Calculate
    calculate_button = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "form#flightForm > button[type='button']"))
    )
    calculate_button.click()

    # Get calculated fare
    calculated_fare = WebDriverWait(setup, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "div#result"))
    ).text

    result = "Pass" if calculated_fare == expected_cost else 'Fail'

    # Write the calculated fare to the Excel file
    write_to_excel(row + 2, calculated_fare,
                   result)  # row + 2 because Excel rows are 1-indexed and there's a header row

    # Assert
    assert result == "Pass", f"Test failed for row {row + 2}. Expected: {expected_cost}, Got: {calculated_fare}"
